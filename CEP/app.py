from flask import Flask, render_template, request, send_file, session
import pickle
import numpy as np
import requests
from datetime import datetime, timedelta, date
import os

# IMPORTANT for Render (no display server)
import matplotlib
matplotlib.use('Agg')

# Excel in memory
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

app = Flask(__name__)
app.secret_key = "smartirrigate_secret"

# Load model
model_path = os.path.join(os.path.dirname(__file__), 'irrigation_model.pkl')
with open(model_path, 'rb') as f:
    model = pickle.load(f)

API_KEY = "4ad5dac7e80eaae2c8fee266fa35043e"


# ================= WEATHER =================

def get_weather(city):
    try:
        url = f"https://api.openweathermap.org/data/2.5/weather?q={city}&appid={API_KEY}&units=metric"
        response = requests.get(url, timeout=5)
        data = response.json()

        temperature = data["main"]["temp"]
        humidity = data["main"]["humidity"]
        rainfall = data.get("rain", {}).get("1h", 0)

    except:
        temperature = 30
        humidity = 60
        rainfall = 0

    return temperature, humidity, rainfall


def get_forecast(city):
    try:
        url = f"https://api.openweathermap.org/data/2.5/forecast?q={city}&appid={API_KEY}&units=metric"
        response = requests.get(url, timeout=5)
        data = response.json()

        forecast_days = []
        for item in data["list"]:
            if "12:00:00" in item["dt_txt"]:
                date_str = item["dt_txt"].split(" ")[0]
                day_name = datetime.strptime(date_str, "%Y-%m-%d").strftime("%A")

                forecast_days.append({
                    "day": day_name,
                    "temp": item["main"]["temp"],
                    "humidity": item["main"]["humidity"],
                    "rainfall": item.get("rain", {}).get("3h", 0)
                })

            if len(forecast_days) == 8:
                break

        return forecast_days

    except:
        return []


# ================= ROUTES =================

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/predict', methods=['POST'])
def predict():
    try:
        location   = request.form['location']
        soil_type  = int(request.form['soil_type'])
        crop_stage = int(request.form['crop_stage'])
        field_size = float(request.form['field_size']) if request.form['field_size'] else 1.0

        temperature, humidity, rainfall = get_weather(location)

        features = np.array([[temperature, humidity, rainfall, soil_type, crop_stage]])
        irrigation_pred = model.predict(features)[0]

        # ===== Water per acre =====
        if crop_stage == 1:
            water_per_acre = 300
        elif crop_stage == 2:
            water_per_acre = 500
        else:
            water_per_acre = 400

        # ===== TODAY =====
        if irrigation_pred == 1 or (humidity < 35 and rainfall < 5):
            alert = "Irrigation Required Today"
            water_needed = round(water_per_acre * field_size, 2)
        elif humidity < 45:
            alert = "Irrigation Recommended Soon"
            water_needed = round((water_per_acre * field_size) * 0.5, 2)
        else:
            alert = "No Irrigation Needed"
            water_needed = 0

        forecast_weather = get_forecast(location)
        today_date = date.today()

        week_prediction = []

        for i in range(7):

            if i < len(forecast_weather):
                day_data = forecast_weather[i]
            else:
                last_day = forecast_weather[-1] if forecast_weather else {
                    "temp": 30, "humidity": 60, "rainfall": 0
                }
                temp = last_day["temp"] + np.random.randint(-2, 3)
                hum = max(0, min(100, last_day["humidity"] + np.random.randint(-5, 6)))
                rain = max(0, last_day["rainfall"] + np.random.randint(0, 3))
                day_data = {"temp": temp, "humidity": hum, "rainfall": rain}

            day_name = (today_date + timedelta(days=i)).strftime("%A")

            temp = day_data["temp"]
            hum  = day_data["humidity"]
            rain = day_data["rainfall"]

            features = np.array([[temp, hum, rain, soil_type, crop_stage]])
            irrigation = model.predict(features)[0]

            if crop_stage == 1:
                base_water = 300
            elif crop_stage == 2:
                base_water = 500
            else:
                base_water = 400

            temp_factor = (temp - 25) * 8
            humidity_factor = hum * 2
            rain_factor = rain * 10

            water = base_water + temp_factor - humidity_factor - rain_factor
            if water < 0:
                water = 0

            water = round(water * field_size, 2)

            if irrigation == 1:
                irrigation_status = "Irrigation Required"
                water = round(water)
            elif irrigation == 0 and hum < 40 and rain < 5:
                irrigation_status = "Irrigation Required"
                water = round(water * 0.5)
            else:
                irrigation_status = "No Irrigation"
                water = 0

            week_prediction.append({
                "day": day_name,
                "irrigation": irrigation_status,
                "water": water
            })


        # ===== GRAPH =====
        import matplotlib.pyplot as plt

        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        static_dir = os.path.join(BASE_DIR, "static")
        os.makedirs(static_dir, exist_ok=True)

        graph_path = os.path.join(static_dir, "water_graph.png")

        days = [d["day"] for d in week_prediction]
        water_values = [d["water"] for d in week_prediction]

        plt.figure()
        plt.plot(days, water_values, marker='o')
        plt.title("6-Day Irrigation Water Requirement")
        plt.xticks(rotation=30)
        plt.grid()
        plt.savefig(graph_path, bbox_inches='tight')
        plt.close()


        # ===== SAVE DATA FOR EXCEL =====
        session['excel_data'] = {
            "alert": alert,
            "water_needed": water_needed,
            "week_prediction": week_prediction,
            "location": location,
            "temperature": temperature,
            "humidity": humidity,
            "rainfall": rainfall,
            "soil_type": soil_type,
            "crop_stage": crop_stage,
        }

        return render_template(
            'result.html',
            location=location,
            soil_type=soil_type,
            crop_stage=crop_stage,
            temperature=temperature,
            humidity=humidity,
            rainfall=rainfall,
            irrigation=irrigation_pred,
            water_needed=water_needed,
            alert=alert,
            week_prediction=week_prediction,
            graph="water_graph.png"
        )

    except Exception as e:
        return f"Error: {str(e)}"


# ================= EXCEL DOWNLOAD =================

# Translation dictionaries
TRANSLATIONS = {
    "en": {
        "title": "Smart Irrigation Report",
        "generated": "Generated on",
        "location": "Location",
        "temperature": "Temperature (°C)",
        "humidity": "Humidity (%)",
        "rainfall": "Rainfall (mm)",
        "soil_type": "Soil Type",
        "crop_stage": "Crop Stage",
        "today_status": "Today's Status",
        "water_today": "Water Needed Today (Litres)",
        "week_plan": "7-Day Irrigation Plan",
        "day": "Day",
        "irrigation": "Irrigation Status",
        "water": "Water Required (Litres)",
        "graph_title": "Water Requirement Graph",
        "soils": {1: "Black Soil", 2: "Clay Soil", 3: "Sandy Soil"},
        "stages": {1: "Seedling", 2: "Middle Stage", 3: "Mature Stage"},
        "irr_required": "Irrigation Required",
        "no_irr": "No Irrigation",
        "days": {
            "Monday": "Monday", "Tuesday": "Tuesday", "Wednesday": "Wednesday",
            "Thursday": "Thursday", "Friday": "Friday", "Saturday": "Saturday", "Sunday": "Sunday"
        }
    },
    "hi": {
        "title": "स्मार्ट सिंचाई रिपोर्ट",
        "generated": "तैयार किया गया",
        "location": "स्थान",
        "temperature": "तापमान (°C)",
        "humidity": "आर्द्रता (%)",
        "rainfall": "वर्षा (mm)",
        "soil_type": "मिट्टी का प्रकार",
        "crop_stage": "फसल अवस्था",
        "today_status": "आज की स्थिति",
        "water_today": "आज आवश्यक पानी (लीटर)",
        "week_plan": "7 दिन की सिंचाई योजना",
        "day": "दिन",
        "irrigation": "सिंचाई स्थिति",
        "water": "आवश्यक पानी (लीटर)",
        "graph_title": "पानी की आवश्यकता ग्राफ",
        "soils": {1: "काली मिट्टी", 2: "चिकनी मिट्टी", 3: "रेतीली मिट्टी"},
        "stages": {1: "अंकुर अवस्था", 2: "मध्य अवस्था", 3: "परिपक्व अवस्था"},
        "irr_required": "सिंचाई आवश्यक",
        "no_irr": "सिंचाई नहीं",
        "days": {
            "Monday": "सोमवार", "Tuesday": "मंगलवार", "Wednesday": "बुधवार",
            "Thursday": "गुरुवार", "Friday": "शुक्रवार", "Saturday": "शनिवार", "Sunday": "रविवार"
        }
    },
    "mr": {
        "title": "स्मार्ट सिंचन अहवाल",
        "generated": "तयार केले",
        "location": "स्थान",
        "temperature": "तापमान (°C)",
        "humidity": "आर्द्रता (%)",
        "rainfall": "पाऊस (mm)",
        "soil_type": "मातीचा प्रकार",
        "crop_stage": "पिकाची अवस्था",
        "today_status": "आजची स्थिती",
        "water_today": "आज लागणारे पाणी (लिटर)",
        "week_plan": "७ दिवसांचे सिंचन नियोजन",
        "day": "दिवस",
        "irrigation": "सिंचन स्थिती",
        "water": "आवश्यक पाणी (लिटर)",
        "graph_title": "पाण्याची गरज ग्राफ",
        "soils": {1: "काळी माती", 2: "चिकण माती", 3: "वालुकामय माती"},
        "stages": {1: "रोप अवस्था", 2: "मध्य अवस्था", 3: "परिपक्व अवस्था"},
        "irr_required": "सिंचन आवश्यक",
        "no_irr": "सिंचन नाही",
        "days": {
            "Monday": "सोमवार", "Tuesday": "मंगळवार", "Wednesday": "बुधवार",
            "Thursday": "गुरुवार", "Friday": "शुक्रवार", "Saturday": "शनिवार", "Sunday": "रविवार"
        }
    }
}

def make_border():
    side = Side(style='thin', color='CCCCCC')
    return Border(left=side, right=side, top=side, bottom=side)

def style_header_cell(cell, bg_color="2E7D32", font_color="FFFFFF"):
    cell.font = Font(name='Arial', bold=True, color=font_color, size=11)
    cell.fill = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = make_border()

def style_data_cell(cell, bold=False, bg_color=None, font_color="000000", align="center"):
    cell.font = Font(name='Arial', bold=bold, color=font_color, size=10)
    if bg_color:
        cell.fill = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = make_border()


@app.route('/download_excel')
def download_excel():
    try:
        data = session.get('excel_data')
        if not data:
            return "Session expired. Please try again.", 400

        lang = request.args.get('lang', 'en')
        if lang not in TRANSLATIONS:
            lang = 'en'
        t = TRANSLATIONS[lang]

        wb = Workbook()
        ws = wb.active
        ws.title = t["title"][:31]

        # Column width
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20

        # ================= TITLE =================
        ws.merge_cells('A1:C1')
        title = ws['A1']
        title.value = t["title"]
        title.font = Font(bold=True, size=14, color="FFFFFF")
        title.fill = PatternFill("solid", start_color="1B5E20")
        title.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30

        # ================= HEADERS =================
        headers = [t["day"], t["irrigation"], t["water"]]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            style_header_cell(cell)

        # ================= TABLE DATA =================
        row_idx = 4

        # TRANSLATE ALERT (IMPORTANT FIX)
        alert_raw = data.get("alert", "")

        alert_map = {
            "Irrigation Required Today": {
                "hi": "आज सिंचाई आवश्यक है",
                "mr": "आज सिंचन आवश्यक आहे"
            },
            "Irrigation Recommended Soon": {
                "hi": "जल्द सिंचाई की सिफारिश",
                "mr": "लवकरच सिंचन शिफारस"
            },
            "No Irrigation Needed": {
                "hi": "सिंचाई की आवश्यकता नहीं",
                "mr": "सिंचनाची गरज नाही"
            }
        }

        if lang != "en" and alert_raw in alert_map:
            alert_display = alert_map[alert_raw][lang]
        else:
            alert_display = alert_raw

        # TODAY LABEL FIX
        today_label_map = {
            "en": "Today",
            "hi": "आज",
            "mr": "आज"
        }

        # TODAY ROW
        dc = ws.cell(row=row_idx, column=1, value=today_label_map.get(lang, "Today"))
        style_data_cell(dc, bg_color="E3F2FD", bold=True)

        ic = ws.cell(row=row_idx, column=2, value=alert_display)
        style_data_cell(ic, bold=True, bg_color="E3F2FD")

        wc = ws.cell(row=row_idx, column=3, value=data["water_needed"])
        style_data_cell(wc, bg_color="E3F2FD")

        row_idx += 1

        # NEXT 6 DAYS
        for day_data in data["week_prediction"][1:]:

            day_en = day_data["day"]
            day_label = t["days"].get(day_en, day_en)

            irr_en = day_data["irrigation"]

            if irr_en == "Irrigation Required":
                irr_label = t["irr_required"]
                row_bg = "FFEBEE"
                irr_color = "C62828"
            else:
                irr_label = t["no_irr"]
                row_bg = "E8F5E9"
                irr_color = "2E7D32"

            dc = ws.cell(row=row_idx, column=1, value=day_label)
            style_data_cell(dc, bg_color=row_bg)

            ic = ws.cell(row=row_idx, column=2, value=irr_label)
            style_data_cell(ic, bold=True, bg_color=row_bg, font_color=irr_color)

            wc = ws.cell(row=row_idx, column=3, value=day_data["water"])
            style_data_cell(wc, bg_color=row_bg)

            row_idx += 1

        # ================= GRAPH =================
        row_idx += 2

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
        gt = ws.cell(row=row_idx, column=1)
        gt.value = t["graph_title"]
        style_header_cell(gt)

        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        graph_path = os.path.join(BASE_DIR, "static", "water_graph.png")

        if os.path.exists(graph_path):
            img = XLImage(graph_path)
            img.width = 500
            img.height = 300
            ws.add_image(img, f"A{row_idx+1}")

        # ================= SAVE =================
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        filename_map = {
            "en": "irrigation_report_en.xlsx",
            "hi": "irrigation_report_hi.xlsx",
            "mr": "irrigation_report_mr.xlsx"
        }

        return send_file(
            file_stream,
            as_attachment=True,
            download_name=filename_map.get(lang, "irrigation_report.xlsx"),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return f"Excel Error: {str(e)}"

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
